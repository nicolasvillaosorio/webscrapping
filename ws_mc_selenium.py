import requests
from PIL import Image as PILImage  
from io import BytesIO
from openpyxl.drawing.image import Image 
from openpyxl import Workbook

from os import remove


from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from time import sleep
import pickle

### inicializacion del webdriver

service = Service(executable_path=r'c:\webdriver\msedgedriver.exe')
driver = webdriver.Edge(service=service)

#driver.get("https://listado.mercadolibre.com.co/xiaomi-watch#D[A:xiaomi%20watch]")
driver.get("https://www.mercadolibre.com/jms/mco/lgz/msl/login")
# session_cookies= driver.get_cookies()
# print(session_cookies)
sleep(5)
########
### inyeccion de cookies
m_cookies= pickle.load(open("cookies.pkl", "rb"))

for cookie in m_cookies:
    cookie["domain"]= ".mercadolibre.com"
    try:
        driver.add_cookie(cookie)

    except:
        pass
    driver.refresh()

# all_cookies = driver.get_cookies()

# print(all_cookies)

### creacion del workbook para el archivo excel
wb= Workbook()
ws = wb.active
ws.title="smartwatches"
ws.append(["Nombre", "Precio", "Link", "Imagen"])

## inicio del scraping

driver.get("https://listado.mercadolibre.com.co/xiaomi-watch#D[A:xiaomi%20watch]")
all =  driver.find_elements(By.CSS_SELECTOR, "li.ui-search-layout__item.shops__layout-item")
sleep(3)

wait = WebDriverWait(driver, 10)
all = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "li.ui-search-layout__item.shops__layout-item")))


for index, item in enumerate(all, start=2):

    relojes={}

    nombre = item.find_element(By.CSS_SELECTOR, "h2.poly-box.poly-component__title").text
    precio= (item.find_element(By.CSS_SELECTOR,"span.andes-money-amount.andes-money-amount--cents-superscript").text)
    precio = precio.replace("\n", "")

    link = item.find_element(By.TAG_NAME, "a").get_attribute("href")

    #print(f"nombre: {link}")
    ws.cell(row=index, column=1, value=nombre)
    ws.cell(row=index, column=2, value=precio)  
    ws.cell(row=index, column=3, value=link)

    img_url = item.find_element(By.CSS_SELECTOR, "img").get_attribute("data-src") or item.find_element(By.CSS_SELECTOR, "img").get_attribute("src")

    ### solicitud get para obtener la imagen del producto
    ### e insercion de la imagen al archivo excel
    full_img = requests.get(img_url)

    if full_img.status_code == 200:
        temp_filename = f"temp_image_{index}.png"
        img_data = BytesIO(full_img.content)

        with PILImage.open(img_data) as img:
            img = img.convert("RGB") 
            img.save(temp_filename) 

        img_aux = Image(temp_filename)
        img_aux.width = 100 
        img_aux.height = 100  
        ws.add_image(img_aux, f"D{index}")

    
    #index+=1

wb.save("smartwatches.xlsx")

### ajuste del ancho de las columnas

for colum in ws.columns:
    long_max = 0
    colum_letra = colum[0].column_letter  
    for cell in colum:
        try:
            if cell.value:  
                long_max = max(long_max, len(str(cell.value)))
        except:
            pass
    if colum_letra == "C":
        long_max/=30 ## como es un enlace web es demasiado largo, 
                    ##por lo que se divide para que quede una linea menos ancha
    ws.column_dimensions[colum_letra].width = long_max + 2

wb.save("smartwatches.xlsx")

##borrado de las imagenes descargadas para el archivo excel
for index, item in enumerate(all, start=2):
    remove(f"temp_image_{index}.png")
